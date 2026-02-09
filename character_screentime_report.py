"""
Generate Excel report from character identification results.
Creates sheets with:
- Overall Screen Time (all characters)
- Individual sheets for top 5 characters with occurrence details
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from argparse import ArgumentParser


def load_character_database(json_path: Path) -> dict:
    """Load character database from JSON."""
    with open(json_path, "r") as f:
        return json.load(f)


def calculate_screen_time(appearance_count: int, frame_interval_sec: float) -> float:
    """Calculate screen time in seconds from frame count."""
    return appearance_count * frame_interval_sec


def create_excel_report(
    character_data: dict,
    output_path: Path,
    video_duration_sec: float,
    frame_interval_sec: float = 1.0,
    top_n: int = 5,
):
    """
    Create Excel report with overall screen time and individual character sheets.
    
    Args:
        character_data: Dict from JSON with actor names as keys
        output_path: Path to save Excel file
        video_duration_sec: Total video duration in seconds
        frame_interval_sec: Frame sampling interval (default 1.0)
        top_n: Number of top characters to create individual sheets for
    """
    
    # Calculate screen times
    character_times = {}
    for actor_name, data in character_data.items():
        screen_time = calculate_screen_time(data["count"], frame_interval_sec)
        character_times[actor_name] = {
            "screen_time": screen_time,
            "count": data["count"],
            "on_screen_names": data["on_screen_names"],
            "appearances": data["appearances"],
        }
    
    # Sort by screen time
    sorted_chars = sorted(
        character_times.items(),
        key=lambda x: x[1]["screen_time"],
        reverse=True
    )
    
    top_characters = sorted_chars[:top_n]
    
    # Create workbook
    wb = Workbook()
    ws_overall = wb.active
    ws_overall.title = "Overall Screen Time"
    
    # --- Overall Screen Time Sheet ---
    print("[1/3] Creating Overall Screen Time sheet...")
    
    # Header
    headers = ["Rank", "Actor Name", "On-Screen Characters", "Screen Time (sec)", "Frame Count", "Percentage (%)"]
    ws_overall.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws_overall[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data
    total_screen_time = sum(t["screen_time"] for _, t in sorted_chars)
    
    for rank, (actor_name, data) in enumerate(sorted_chars, 1):
        screen_time = data["screen_time"]
        count = data["count"]
        on_screen = ", ".join(data["on_screen_names"])
        pct = (screen_time / total_screen_time * 100) if total_screen_time > 0 else 0
        
        ws_overall.append([rank, actor_name, on_screen, round(screen_time, 1), count, round(pct, 1)])
    
    # Adjust column widths
    ws_overall.column_dimensions["A"].width = 6
    ws_overall.column_dimensions["B"].width = 30
    ws_overall.column_dimensions["C"].width = 50
    ws_overall.column_dimensions["D"].width = 18
    ws_overall.column_dimensions["E"].width = 12
    ws_overall.column_dimensions["F"].width = 14
    
    # Center align numeric columns
    for row in ws_overall.iter_rows(min_row=2, max_row=ws_overall.max_row, min_col=1, max_col=6):
        for cell in row:
            if cell.column in [1, 4, 5, 6]:
                cell.alignment = Alignment(horizontal="center")
    
    # --- Individual Character Sheets ---
    print("[2/3] Creating individual character sheets...")
    
    for rank, (actor_name, data) in enumerate(top_characters, 1):
        screen_time = data["screen_time"]
        count = data["count"]
        on_screen = ", ".join(data["on_screen_names"])
        appearances = data["appearances"]
        
        # Create sheet
        safe_name = actor_name.replace("/", "-").replace("\\", "-")[:31]
        ws_char = wb.create_sheet(title=f"{rank}. {safe_name}")
        
        # Header with actor info
        ws_char.append([f"Character: {actor_name}"])
        ws_char.append([f"On-Screen Names/Roles: {on_screen}"])
        ws_char.append([f"Total Screen Time: {round(screen_time, 1)} seconds"])
        ws_char.append([f"Appearance Count: {count}"])
        ws_char.append([])  # Blank row
        
        # Style title rows
        title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        title_font = Font(bold=True)
        for i in range(1, 5):
            for cell in ws_char[i]:
                cell.fill = title_fill
                cell.font = title_font
        
        # Occurrences table
        ws_char.append(["Occurrence #", "Start Time (sec)", "Duration (sec)"])
        
        # Style occurrence header
        occ_header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        occ_header_font = Font(bold=True, color="FFFFFF")
        for cell in ws_char[6]:
            cell.fill = occ_header_fill
            cell.font = occ_header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Group consecutive appearances into segments
        if appearances:
            appearances_sorted = sorted(appearances)
            segments = []
            current_start = appearances_sorted[0]
            current_end = appearances_sorted[0]
            
            for t in appearances_sorted[1:]:
                # If within 2 seconds, consider it the same segment
                if t - current_end <= 2.0:
                    current_end = t
                else:
                    segments.append((current_start, current_end + frame_interval_sec))
                    current_start = t
                    current_end = t
            
            # Add last segment
            segments.append((current_start, current_end + frame_interval_sec))
            
            # Add segment data
            for occ_idx, (start, end) in enumerate(segments, 1):
                duration = round(end - start, 1)
                ws_char.append([occ_idx, round(start, 1), duration])
            
            # Adjust column widths
            ws_char.column_dimensions["A"].width = 14
            ws_char.column_dimensions["B"].width = 18
            ws_char.column_dimensions["C"].width = 16
            
            # Center align
            for row in ws_char.iter_rows(min_row=7, max_row=ws_char.max_row, min_col=1, max_col=3):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
    
    # Return workbook for further modifications
    return wb


def add_dialogues_sheet(wb, dialogues_data: dict):
    """
    Add a sheet with top 5 punchy dialogues to existing workbook.
    
    Args:
        wb: Existing openpyxl Workbook object
        dialogues_data: Dict from punchy_dialogues.py with dialogue info
    """
    
    if not dialogues_data:
        return wb
    
    dialogues = dialogues_data.get("validated_dialogues", [])
    if not dialogues:
        return wb
    
    # Create new sheet
    ws_dialogues = wb.create_sheet("Top 5 Dialogues")
    
    # Header
    header_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ["Rank", "Dialogue", "Time (sec)", "Speaker", "Context", "Reasoning"]
    ws_dialogues.append(headers)
    
    # Style header
    for cell in ws_dialogues[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    
    # Add dialogue data
    for idx, dialogue in enumerate(dialogues[:5], 1):
        ws_dialogues.append([
            idx,
            dialogue.get("dialogue", ""),  # Full dialogue text
            round(dialogue.get("start_second", 0), 1),
            dialogue.get("speaker", "Unknown"),
            dialogue.get("context", ""),  # Full context (no truncation)
            dialogue.get("reasoning", "")
        ])
    
    # Adjust column widths and enable text wrapping
    ws_dialogues.column_dimensions["A"].width = 8
    ws_dialogues.column_dimensions["B"].width = 50
    ws_dialogues.column_dimensions["C"].width = 12
    ws_dialogues.column_dimensions["D"].width = 20
    ws_dialogues.column_dimensions["E"].width = 60  # Increased for full context
    ws_dialogues.column_dimensions["F"].width = 70  # Increased for full reasoning
    
    # Center align rank and time
    for row in ws_dialogues.iter_rows(min_row=2, max_row=ws_dialogues.max_row):
        row[0].alignment = Alignment(horizontal="center")
        row[2].alignment = Alignment(horizontal="center")
        row[1].alignment = Alignment(wrap_text=True)
        row[4].alignment = Alignment(wrap_text=True)
        row[5].alignment = Alignment(wrap_text=True)
    
    return wb


def save_excel_workbook(wb, output_path: Path):
    """Save workbook to file."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = ArgumentParser(description="Generate Excel report from character identification results")
    parser.add_argument("character_json", help="Path to character database JSON file (from character_identification.py)")
    parser.add_argument("--output", help="Output Excel file path (default: same name as JSON but .xlsx)")
    parser.add_argument("--video-duration", type=float, help="Video duration in seconds (optional)")
    parser.add_argument("--interval", type=float, default=1.0, help="Frame sampling interval in seconds (default: 1.0)")
    parser.add_argument("--top-n", type=int, default=5, help="Number of top characters to create detailed sheets for (default: 5)")
    
    args = parser.parse_args()
    
    json_path = Path(args.character_json)
    if not json_path.exists():
        raise FileNotFoundError(f"Character database not found: {json_path}")
    
    # Determine output path
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = json_path.parent / f"{json_path.stem}_screentime.xlsx"
    
    # Load character data
    character_data = load_character_database(json_path)
    
    # Estimate video duration if not provided
    if args.video_duration:
        video_duration = args.video_duration
    else:
        # Estimate from last appearance time + some buffer
        all_times = []
        for data in character_data.values():
            all_times.extend(data["appearances"])
        video_duration = max(all_times) + 10 if all_times else 0
        print(f"Estimated video duration: {round(video_duration, 1)} seconds")
    
    # Create report
    print(f"Loading character data from: {json_path}")
    create_excel_report(
        character_data,
        output_path,
        video_duration_sec=video_duration,
        frame_interval_sec=args.interval,
        top_n=args.top_n,
    )


if __name__ == "__main__":
    main()
